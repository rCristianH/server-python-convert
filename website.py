import csv
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Border, Side, Font, Alignment
from flask import Flask, request, render_template, send_file

app = Flask(__name__)


@app.route("/", methods=["GET", "POST"])
def formulario():
    if request.method == "POST":
        csv_file = request.files["csv_file"]
        mes = request.form["mes"]
        saldo_inicial = float(request.form["saldo_inicial"])
        csv_temp_file = "temp.csv"
        csv_file.save(csv_temp_file)
        # Abre el archivo CSV
        with open(csv_temp_file, newline='', encoding='utf-8') as csvfile:
            reader = csv.DictReader(csvfile)
            #consulta el mes
            segunda_fila = next(reader, None)
            fecha = segunda_fila['Fecha']   
            partes_fecha = fecha.split('-')
            numero_mes = partes_fecha[1]
            ficha = numero_mes + " de 12"

            # Inicializa variables
            nit = "Nit. 94471472-5"
            saldo = saldo_inicial
            # Diccionarios para almacenar los valores de ingreso, compras y gastos por día
            ingresos_por_dia = {}
            compras_por_dia = {}
            gastos_por_dia = {}

            # Procesa las filas del CSV
            for row in reader:
                tipo = row["Tipo"]
                fecha = row["Fecha"]
                total = float(row["Total"])

                # Calcula los valores de ingreso, compras y gastos
                if tipo == "Factura":
                    if fecha not in ingresos_por_dia:
                        ingresos_por_dia[fecha] = 0
                    ingresos_por_dia[fecha] += total
                elif tipo == "Factura proveedor":
                    if fecha not in compras_por_dia:
                        compras_por_dia[fecha] = 0
                    compras_por_dia[fecha] += total
                elif tipo == "Pago varios":
                    if fecha not in gastos_por_dia:
                        gastos_por_dia[fecha] = 0
                    gastos_por_dia[fecha] += total

            # Crea un nuevo archivo Excel
            wb = Workbook()
            ws = wb.active



            def modificar_ancho(i,j):
                ancho_columna = i  # Cambia este valor según tus 
                ws.column_dimensions[j].width = ancho_columna

            modificar_ancho(17,"A")
            modificar_ancho(17,"B")
            modificar_ancho(17,"C")
            modificar_ancho(17,"D")
            modificar_ancho(17,"E")


            
            """ partes_fecha = fecha.split('-')
            numero_mes = partes_fecha[1] """

            # Define las columnas en el archivo Excel
            ws.append(["LIBRO DE REGISTRO DE OPERACIONES DIARIAS FOLIO No" ])
            ws.append(["NOMBRE", "Pinturas del centro"])
            ws.append(["IDENTIFICACIÓN", nit])
            ws.append(["MES", mes])
            ws.append(["SALDO INICIAL", saldo_inicial])
            ws.append(["FECHA", "INGRESO", "COMPRAS", "GASTOS", "SALDO"])

            # Procesa los días y calcula los saldos
            for fecha in sorted(set(ingresos_por_dia.keys()) | set(compras_por_dia.keys()) | set(gastos_por_dia.keys())):
                ingreso_diario = ingresos_por_dia.get(fecha, 0)
                compras_diario = compras_por_dia.get(fecha, 0)
                gastos_diario = gastos_por_dia.get(fecha, 0)

                saldo += ingreso_diario - compras_diario - gastos_diario
                ws.append([fecha, ingreso_diario,
                          compras_diario, gastos_diario, saldo])

            # Calcula y agrega los totales
            total_ingresos = sum(ingresos_por_dia.values())
            total_compras = sum(compras_por_dia.values())
            total_gastos = sum(gastos_por_dia.values())
            ws.append(["Total", total_ingresos, total_compras, total_gastos])

            # Establece el estilo de fuente en negrita para los títulos
            font_bold = Font(bold=True)

            def aplicar_formato(cell, alt, row, sc, ec):
                ws[cell].font = font_bold
                mes_celda = ws[alt]
                mes_celda.alignment = Alignment(
                    horizontal='center', vertical='center')
                ws.merge_cells(start_row=row, start_column=sc,
                               end_row=row, end_column=ec)
            aplicar_formato("A1", "A1", 1, 1, 4)
            aplicar_formato("A2", "B2", 2, 2, 5)
            aplicar_formato("A3", "B3", 3, 2, 5)
            aplicar_formato("A4", "B4", 4, 2, 5)
            aplicar_formato("A5", "B5", 5, 2, 5)
            ws['E1'] = ficha
            ws['E1'].alignment = Alignment(
                    horizontal='center', vertical='center')
            
            def aplicar_formato_contabilidad():
                # Aplica el estilo de fuente en negrita a los títulos
                for cell in ws[6]:  # Suponiendo que los títulos están en la primera fila
                    cell.font = font_bold

                # Encuentra el final de la tabla buscando datos en una columna específica (por ejemplo, columna B)
                columna_busqueda = ws['B']
                tabla_fin = None

                for cell in reversed(columna_busqueda):
                    if cell.value is not None:
                        tabla_fin = cell.row
                        break

                # Obtén el rango de celdas de la tabla
                if tabla_fin is not None:
                    tabla_rango = ws["B7:E{}".format(tabla_fin)]
                else:
                    # Manejo de caso donde no se encuentra el final de la tabla
                    # Fallback a una sola fila (puedes ajustarlo según tus necesidades)
                    tabla_rango = ws["B7:E7"]

                # Aplicar estilo de contabilidad a peso colombiano con 2 decimales
                contabilidad_style = NamedStyle(name="contabilidad")
                contabilidad_style.number_format = '_("$"* #,##0.00_);_("$"* -#,##0.00_);_("$"* "-"??_);_(@_)'

                # Aplica el estilo de contabilidad al rango de celdas
                for row in tabla_rango:
                    for cell in row:
                        cell.style = contabilidad_style
                for fila in ws.iter_rows(min_row=4, max_row=4, min_col=2, max_col=5):
                    for cell in fila:
                        cell.style = contabilidad_style
            aplicar_formato_contabilidad()
            


            def encontrar_ultima_fila(columna):
                # Carga el archivo Excel

                # Obtiene la columna específica
                columna_busqueda = ws[columna]

                # Busca la última fila con datos en la columna
                for fila, cell in enumerate(reversed(columna_busqueda), start=1):
                    if cell.value is not None:
                        return cell.row

                # Si no se encuentran datos en la columna, devuelve 0
                return 7
            celda_total = f"A{encontrar_ultima_fila('A')}"
            ws[celda_total].font = font_bold
            # Definir el rango de celdas que contiene datos en la tabla
            inicio_fila = 1  # Cambia esto según la ubicación real de tus datos
            fin_columna = ws.max_column
            fin_fila = ws.max_row
            #centra
            ws['B4'].alignment = Alignment(
                    horizontal='center', vertical='center')
            # Crea un objeto de estilo de borde en negrita
            border = Border(left=Side(style='thick'), right=Side(
                style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))

            # Aplica el borde en negrita a las celdas de la tabla
            for fila in ws.iter_rows(min_row=inicio_fila, max_row=fin_fila, min_col=1, max_col=fin_columna):
                for cell in fila:
                    cell.border = border

            # Guarda el archivo Excel
            wb.save('resultado.xlsx')

        return send_file('resultado.xlsx', as_attachment=True, download_name="archivo_excel.xlsx")

    return render_template("formulario.html")


if __name__ == "__main__":
    app.run(debug=True)


print("Archivo Excel generado con éxito.")
